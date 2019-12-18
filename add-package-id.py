#!/usr/bin/env python
# encoding: utf-8

import os
import csv
import glob

from openpyxl import load_workbook
from shutil import copy

class DataFile(object):

    def __init__(self, filename):
        self.delimiter = '.'
        self.t_list = [".xlsx", ".csv"]

        self.filename = filename
        self.prefix, self.filetype = self._getFileType()

    def _getFileType(self):
        prefix, suffix = os.path.splitext(self.filename)

        # check filename
        if suffix in self.t_list:
            return (prefix, suffix)
        else:
            raise Exception("Filename format ERROR!")

    def _checkFileExists(self):
        if not os.path.exists(self.filename):
            raise Exception("File not found!")

class ShipDataFile(DataFile):

    #[订单标识，商品交易号]
    def read(self):
        if self.filetype != ".csv":
            raise Exception("DataFile type must be 'ship'!")

        self._checkFileExists()

        data = list()
        with open(self.filename, 'r+', newline='', encoding='GBK') as f:
            source = csv.reader(f)
            for r in source:
                data.append(r[:2])
        return dict(data[1:])

    def write(self, data):
        with open(self.filename, 'w+', encoding='GBK') as f:
            w = csv.writer(f)
            w.writerows(data)

class ExpressDataFile(DataFile):

    def read(self):
        if self.filetype != ".xlsx":
            raise Exception("DataFile type must be 'express'!")

        self._checkFileExists()

        self.workbook = load_workbook(filename = self.filename)

        return self.workbook

    def write(self):
        self.workbook.save(self.filename)

class RecordDataFile(DataFile):

    def __init__(self, filename):
        self.filename = filename

    def write(self, data):
        #print(data)
        with open(self.filename, 'w+', encoding='GBK') as f:
            w = csv.writer(f)
            w.writerows(data)

if __name__ == "__main__":

    # 获取文件名
    # 自动获取文件名
    # ship_filename
    ship_filename = list()
    for i in glob.glob("*.csv"):
        ship_filename.append(i)

    if len(ship_filename) == 1:
        ship_filename = ship_filename[0]
    else:
        raise Exception("*.csv NOT FOUND!")
    #print(ship_filename)
    #exit()

    # express_filename
    express_filename = list()
    for i in glob.glob("*.xlsx"):
        express_filename.append(i)

    if len(express_filename) == 1:
        express_filename = express_filename[0]
    else:
        raise Exception("*xlsx NOT FOUND!")
    #print(express_filename)
    #exit()

    backup_express_filename = "copy_" + express_filename # copy_ + express_filename
    only_in_ship_filename = "only_in_" # only_in_*.txt
    only_in_express_filename = "only_in_" # only_in_*.txt
    #print(backup_express_filename)
    #print(only_in_ship_filename)
    #print(only_in_express_filename)
    #exit()

    # 读取文件
    # ship file
    ship_file = ShipDataFile(ship_filename)
    ship_file_data = ship_file.read() # {"订单标识":"商品交易号"}
    #print(ship_file_data)
    #exit()

    #express file
    express_file = ExpressDataFile(express_filename)
    ws = express_file.read().active # get active worksheet
    #print(ws["A1"].value)
    #exit()

    # 备份原文件
    copy(express_filename, backup_express_filename)

    # 查找修改数据
    interval = 15
    i = 1
    used_ship_data = set()
    unfound_express_data = list()
    while ws["A%d" % (i)].value != None:
        k = str(ws["A%d" % (i)].value)
        if k in ship_file_data.keys():
            used_ship_data.add(k)
            ws["E%d" % (i)] = ship_file_data[k] # add tracking no
        else:
            unfound_express_data.append(k)
        i += interval

    # 保存文件
    express_file.write()

    # 写入记录
    #print(used_ship_data)
    #print(ship_file_data.keys())
    #print(unfound_express_data)
    #exit()

    only_in_ship_filename += ship_file.prefix + ".txt"
    #print(only_in_ship_filename)
    only_in_ship_file = RecordDataFile(only_in_ship_filename)
    #print(set(ship_file_data.keys()).difference(used_ship_data))
    data = [[x] for x in list(set(ship_file_data.keys()).difference(used_ship_data))]
    #print(data)
    only_in_ship_file.write(data)

    only_in_express_filename += express_file.prefix + ".txt"
    #print(only_in_express_filename)
    only_in_express_file = RecordDataFile(only_in_express_filename)
    data = [[x] for x in unfound_express_data]
    #print(data)
    only_in_express_file.write(data)
